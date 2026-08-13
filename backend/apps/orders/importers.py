import io
import openpyxl
from decimal import Decimal
from .models import Order, OrderItem, ExchangeRate, calc_order_profit
from apps.basic_info.models import Customer, Product, Factory

ALI_STATUS_MAP = {
    '待确认': '接单', '待发货': '排产', '已发货': '发货', '交易成功': '签收',
    '交易失败': '已取消', '已退款': '已取消',
}
CANCELLED_STATUSES = {'交易失败', '已退款'}

ORDER_COLUMNS = [
    '订单状态', '产品清单', '订单日期', '联系人', '订单名称', '运费', '物流保险费', '附加费用',
    '订单金额（USD）', '交易服务费(USD)', '运输成本', '承运商', '物流', '单号', '备注',
]
ITEM_COLUMNS = [
    '序号', '供应商', '数量', '型号', '产品规格', '单价', '金额小计', '含税成本价',
    '产品毛利', '毛利润率', '产品编号',
]


def _match_customer(name):
    if not name:
        return None
    name = str(name).strip()
    c = Customer.objects.filter(name=name).first()
    if c:
        return c
    c = Customer.objects.filter(name__iexact=name).first()
    return c


def _match_factory(name):
    if not name:
        return None
    name = str(name).strip()
    f = Factory.objects.filter(name=name).first()
    if f:
        return f
    f = Factory.objects.filter(alias=name).first()
    return f


def _match_product(no):
    if not no:
        return None
    return Product.objects.filter(product_no=str(no).strip()).first()


def import_orders(file):
    wb = openpyxl.load_workbook(file, data_only=True)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    col = {h: i for i, h in enumerate(headers) if h}
    success, failures = 0, []
    unmatched = {'customers': [], 'products': [], 'factories': []}
    # 按 order_no 分组（主表字段可能合并单元格，从每行读但按订单号去重）
    orders_data = {}
    order_row_map = {}
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        order_no = row[col['订单名称']] if '订单名称' in col else None
        if not order_no:
            failures.append({'row': row_idx, 'reason': '订单号为空'})
            continue
        order_no = str(order_no).strip()
        if order_no not in orders_data:
            orders_data[order_no] = {
                'ali_status': row[col['订单状态']] if '订单状态' in col else '',
                'order_date': row[col['订单日期']] if '订单日期' in col else None,
                'contact': row[col['联系人']] if '联系人' in col else '',
                'freight': row[col['运费']] if '运费' in col else 0,
                'insurance': row[col['物流保险费']] if '物流保险费' in col else 0,
                'surcharge': row[col['附加费用']] if '附加费用' in col else 0,
                'amount': row[col['订单金额（USD）']] if '订单金额（USD）' in col else 0,
                'service_fee': row[col['交易服务费(USD)']] if '交易服务费(USD)' in col else 0,
                'transport': row[col['运输成本']] if '运输成本' in col else 0,
                'carrier': row[col['承运商']] if '承运商' in col else '',
                'logistics': row[col['物流']] if '物流' in col else '',
                'tracking_no': row[col['单号']] if '单号' in col else '',
                'remark': row[col['备注']] if '备注' in col else '',
                'row': row_idx,
                'items': [],
            }
            order_row_map[order_no] = row_idx
        # 产品行
        item = {
            'seq': row[col['序号']] if '序号' in col else 0,
            'supplier': row[col['供应商']] if '供应商' in col else '',
            'qty': row[col['数量']] if '数量' in col else 0,
            'model': row[col['型号']] if '型号' in col else '',
            'spec': row[col['产品规格']] if '产品规格' in col else '',
            'price': row[col['单价']] if '单价' in col else 0,
            'subtotal': row[col['金额小计']] if '金额小计' in col else 0,
            'cost': row[col['含税成本价']] if '含税成本价' in col else 0,
            'product_no': row[col['产品编号']] if '产品编号' in col else '',
        }
        orders_data[order_no]['items'].append(item)

    for order_no, d in orders_data.items():
        try:
            customer = _match_customer(d['contact'])
            if d['contact'] and not customer:
                unmatched['customers'].append({'row': d['row'], 'name': str(d['contact'])})
            ali_status = str(d['ali_status'] or '').strip()
            tracking_status = ALI_STATUS_MAP.get(ali_status, '')
            is_cancelled = ali_status in CANCELLED_STATUSES
            order, created = Order.objects.update_or_create(
                order_no=order_no,
                defaults={
                    'ali_status': ali_status, 'tracking_status': tracking_status, 'is_cancelled': is_cancelled,
                    'order_date': d['order_date'],
                    'customer': customer,
                    'salesman': customer.salesman if customer else None,
                    'amount_usd': d['amount'] or 0, 'freight': d['freight'] or 0, 'insurance': d['insurance'] or 0,
                    'surcharge': d['surcharge'] or 0, 'service_fee_usd': d['service_fee'] or 0,
                    'transport_cost': d['transport'] or 0, 'carrier': d['carrier'] or '',
                    'logistics_method': d['logistics'] or '',
                    'tracking_no': d['tracking_no'] or '', 'remark': d['remark'] or '',
                },
            )
            # tracker：新订单填 customer.tracker；已存在若空才填
            if not order.tracker and customer and customer.tracker:
                order.tracker = customer.tracker
                order.save(update_fields=['tracker'])
            # 产品行整组替换
            order.items.all().delete()
            for it in d['items']:
                product = _match_product(it['product_no'])
                if it['product_no'] and not product:
                    unmatched['products'].append({'row': d['row'], 'product_no': str(it['product_no'])})
                factory = _match_factory(it['supplier'])
                if it['supplier'] and not factory:
                    unmatched['factories'].append({'row': d['row'], 'name': str(it['supplier'])})
                OrderItem.objects.create(
                    order=order, seq=it['seq'] or 0, product=product, factory=factory,
                    model=it['model'] or '', product_no=str(it['product_no'] or ''), spec=it['spec'] or '',
                    qty=it['qty'] or 0, unit_price=it['price'] or 0, subtotal=it['subtotal'] or 0,
                    cost_price=it['cost'] or 0,
                )
            calc_order_profit(order)
            success += 1
        except Exception as e:
            failures.append({'row': d['row'], 'reason': str(e)})
    return {'success_count': success, 'fail_count': len(failures), 'failures': failures, 'unmatched': unmatched}


def build_order_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(ORDER_COLUMNS + ITEM_COLUMNS)
    ws.append([
        '待确认', '', '2026-05-12', '吴芳', '示例订单', 100, 10, 5, 1000, 30, 50,
        '圆通', 'EMS', 'T1', '', '1', '华鑫', 10, 'M1', '尺寸:54*35',
        '100', '1000', '720', '', '', 'P001',
    ])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf