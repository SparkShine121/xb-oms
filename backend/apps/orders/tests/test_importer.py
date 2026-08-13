import pytest
from io import BytesIO
import openpyxl
from apps.orders.importers import import_orders, build_order_template, ALI_STATUS_MAP
from apps.orders.models import Order
from apps.basic_info.models import Customer, Product, Factory, Category
from datetime import date


@pytest.fixture
def rate(db):
    from apps.orders.models import ExchangeRate
    return ExchangeRate.objects.create(currency_pair='USD/CNY', rate='7.20', effective_date=date(2026, 1, 1))


def make_xlsx(records):
    # records: list of dict {订单级字段 + items:[产品行字段]}
    wb = openpyxl.Workbook()
    ws = wb.active
    headers = [
        '订单状态', '产品清单', '订单日期', '联系人', '订单名称', '运费', '物流保险费', '附加费用',
        '订单金额（USD）', '交易服务费(USD)', '运输成本', '承运商', '物流', '单号', '备注',
        '序号', '供应商', '数量', '型号', '产品规格', '单价', '金额小计', '含税成本价',
        '产品毛利', '毛利润率', '产品编号',
    ]
    ws.append(headers)
    for rec in records:
        for i, it in enumerate(rec['items']):
            row = [
                rec['ali_status'], '', rec['order_date'], rec['contact'], rec['order_no'],
                rec['freight'], rec['insurance'], rec['surcharge'], rec['amount'],
                rec['service_fee'], rec['transport'], rec['carrier'], rec['logistics'],
                rec['tracking_no'], rec['remark'],
                it['seq'], it['supplier'], it['qty'], it['model'], it['spec'],
                it['price'], it['subtotal'], it['cost'], it.get('profit', ''),
                it.get('profit_rate', ''), it['product_no'],
            ]
            ws.append(row)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def test_import_creates_orders(db, rate):
    Customer.objects.create(name='吴芳')
    buf = make_xlsx([{
        'ali_status': '待确认', 'order_date': '2026-05-12', 'contact': '吴芳', 'order_no': 'O1',
        'freight': 100, 'insurance': 10, 'surcharge': 5, 'amount': 1000, 'service_fee': 30,
        'transport': 50, 'carrier': '圆通', 'logistics': 'EMS', 'tracking_no': 'T1', 'remark': '',
        'items': [{'seq': 1, 'supplier': '华鑫', 'qty': 10, 'model': 'M1', 'spec': 'S',
                    'price': 100, 'subtotal': 1000, 'cost': 720, 'product_no': 'P1'}],
    }])
    r = import_orders(buf)
    assert r['success_count'] == 1 and r['fail_count'] == 0
    o = Order.objects.get(order_no='O1')
    assert o.tracking_status == '接单' and o.is_cancelled is False
    assert o.items.count() == 1
    # customer 关联
    assert o.customer is not None and o.customer.name == '吴芳'


def test_import_cancelled_status(db, rate):
    buf = make_xlsx([{
        'ali_status': '交易失败', 'order_date': '2026-05-12', 'contact': 'X', 'order_no': 'O2',
        'freight': 0, 'insurance': 0, 'surcharge': 0, 'amount': 0, 'service_fee': 0,
        'transport': 0, 'carrier': '', 'logistics': '', 'tracking_no': '', 'remark': '',
        'items': [{'seq': 1, 'supplier': '', 'qty': 1, 'model': '', 'spec': '',
                    'price': 0, 'subtotal': 0, 'cost': 0, 'product_no': ''}],
    }])
    import_orders(buf)
    o = Order.objects.get(order_no='O2')
    assert o.is_cancelled is True and o.tracking_status == '已取消'


def test_import_upsert_replaces_items(db, rate):
    Customer.objects.create(name='吴芳')
    buf1 = make_xlsx([{
        'ali_status': '待确认', 'order_date': '2026-05-12', 'contact': '吴芳', 'order_no': 'O1',
        'freight': 0, 'insurance': 0, 'surcharge': 0, 'amount': 100, 'service_fee': 0,
        'transport': 0, 'carrier': '', 'logistics': '', 'tracking_no': '', 'remark': '',
        'items': [{'seq': 1, 'supplier': '', 'qty': 1, 'model': '', 'spec': '',
                    'price': 100, 'subtotal': 100, 'cost': 72, 'product_no': 'P1'}],
    }])
    import_orders(buf1)
    buf2 = make_xlsx([{
        'ali_status': '待发货', 'order_date': '2026-05-12', 'contact': '吴芳', 'order_no': 'O1',
        'freight': 0, 'insurance': 0, 'surcharge': 0, 'amount': 200, 'service_fee': 0,
        'transport': 0, 'carrier': '', 'logistics': '', 'tracking_no': '', 'remark': '',
        'items': [
            {'seq': 1, 'supplier': '', 'qty': 2, 'model': '', 'spec': '',
             'price': 100, 'subtotal': 200, 'cost': 72, 'product_no': 'P1'},
            {'seq': 2, 'supplier': '', 'qty': 1, 'model': '', 'spec': '',
             'price': 50, 'subtotal': 50, 'cost': 36, 'product_no': 'P2'},
        ],
    }])
    r = import_orders(buf2)
    o = Order.objects.get(order_no='O1')
    assert o.tracking_status == '排产' and str(o.amount_usd) == '200.00'
    assert o.items.count() == 2  # 整组替换


def test_import_unmatched(db, rate):
    buf = make_xlsx([{
        'ali_status': '待确认', 'order_date': '2026-05-12', 'contact': '新客户', 'order_no': 'O3',
        'freight': 0, 'insurance': 0, 'surcharge': 0, 'amount': 100, 'service_fee': 0,
        'transport': 0, 'carrier': '', 'logistics': '', 'tracking_no': '', 'remark': '',
        'items': [{'seq': 1, 'supplier': '新工厂', 'qty': 1, 'model': '', 'spec': '',
                    'price': 100, 'subtotal': 100, 'cost': 72, 'product_no': 'NEWP'}],
    }])
    r = import_orders(buf)
    o = Order.objects.get(order_no='O3')
    assert o.customer is None  # 未匹配
    assert any(u['name'] == '新客户' for u in r['unmatched']['customers'])
    assert any(u['name'] == '新工厂' for u in r['unmatched']['factories'])


def test_import_tracker_auto(db, rate):
    from django.contrib.auth.models import User
    t = User.objects.create_user('tracker1', password='pw123456')
    Customer.objects.create(name='吴芳', tracker=t)
    buf = make_xlsx([{
        'ali_status': '待确认', 'order_date': '2026-05-12', 'contact': '吴芳', 'order_no': 'O1',
        'freight': 0, 'insurance': 0, 'surcharge': 0, 'amount': 100, 'service_fee': 0,
        'transport': 0, 'carrier': '', 'logistics': '', 'tracking_no': '', 'remark': '',
        'items': [{'seq': 1, 'supplier': '', 'qty': 1, 'model': '', 'spec': '',
                    'price': 100, 'subtotal': 100, 'cost': 72, 'product_no': ''}],
    }])
    import_orders(buf)
    assert Order.objects.get(order_no='O1').tracker == t


def test_import_template_download(db):
    buf = build_order_template()
    wb = openpyxl.load_workbook(buf)
    assert wb.active.max_row >= 2  # 表头+示例


def test_import_bad_file(db):
    with pytest.raises(Exception):
        import_orders(BytesIO(b'not an xlsx'))