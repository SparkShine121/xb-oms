import openpyxl
from .models import Product, Factory

PRODUCT_COLUMNS = ['product_no', 'model', 'name', 'spec', 'default_price', 'default_cost_price']

def import_products(file):
    wb = openpyxl.load_workbook(file, data_only=True)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    col_index = {name: i for i, name in enumerate(headers)}
    success, failures = 0, []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        try:
            data = {name: row[col_index[name]] for name in PRODUCT_COLUMNS if name in col_index}
            product_no = (data.get('product_no') or '').strip()
            if not product_no:
                raise ValueError('product_no 不能为空')
            obj, created = Product.objects.update_or_create(
                product_no=product_no,
                defaults={
                    'model': data.get('model') or '',
                    'name': data.get('name') or '',
                    'spec': data.get('spec') or '',
                    'default_price': data.get('default_price') or 0,
                    'default_cost_price': data.get('default_cost_price') or 0,
                },
            )
            success += 1
        except Exception as e:
            failures.append({'row': row_idx, 'reason': str(e)})
    return {'success_count': success, 'fail_count': len(failures), 'failures': failures}

def build_product_template():
    wb = openpyxl.Workbook(); ws = wb.active
    ws.append(PRODUCT_COLUMNS)
    ws.append(['P001', 'M1', '示例名片', '90x54mm', '0.10', '0.50'])
    import io
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf

FACTORY_COLUMNS = ['name', 'alias', 'contact', 'phone', 'settle_currency']

def import_factories(file):
    wb = openpyxl.load_workbook(file, data_only=True)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    col_index = {name: i for i, name in enumerate(headers)}
    success, failures = 0, []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        try:
            data = {name: row[col_index[name]] for name in FACTORY_COLUMNS if name in col_index}
            name = (data.get('name') or '').strip()
            if not name:
                raise ValueError('name 不能为空')
            Factory.objects.update_or_create(
                name=name,
                defaults={
                    'alias': data.get('alias') or '',
                    'contact': data.get('contact') or '',
                    'phone': data.get('phone') or '',
                    'settle_currency': data.get('settle_currency') or 'CNY',
                },
            )
            success += 1
        except Exception as e:
            failures.append({'row': row_idx, 'reason': str(e)})
    return {'success_count': success, 'fail_count': len(failures), 'failures': failures}

def build_factory_template():
    wb = openpyxl.Workbook(); ws = wb.active
    ws.append(FACTORY_COLUMNS)
    ws.append(['华鑫', '外部YT', '张三', '13800000000', 'CNY'])
    import io
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf
